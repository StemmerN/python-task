import argparse
import logging
import os
import sys
import requests
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

# Target URL for the CSV upload
target_url = 'http://127.0.0.1:8003/upload-csv/'

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


def upload_csv(file_path, verbose=False):
    if verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    if not os.path.exists(file_path):
        logging.error('The file does not exist.')
        sys.exit(2)  # Exit code 2 for file not found

    try:
        with open(file_path, 'rb') as file_data:
            files = {'file': file_data}
            logging.info('Starting to upload the file...')
            response = requests.post(target_url, files=files)
            response.raise_for_status()  # Raise HTTP error if status code is not 200
            logging.info('File successfully uploaded.')
            logging.debug(f'Status Code: {response.status_code}')
            logging.debug(f'Response: {response.text}')
            return response.json()  # Return JSON response
    except requests.exceptions.HTTPError as e:
        logging.exception(
            f'HTTP error while uploading the file: {e.response.text if e.response else "No response received"}')
        logging.error(f'Error details: {e.response.content if e.response else "No response received"}')
        logging.error(
            f'Error details (JSON): {e.response.json() if e.response and e.response.content else "No response received"}')
        sys.exit(3)  # Exit code 3 for network or HTTP error
    except Exception as e:
        logging.exception('An unexpected error occurred')
        sys.exit(1)  # Exit code 1 for general errors


def process_and_save_to_excel(response_data, keys, colored, excel_output_path=None):
    keys = list(set(keys))

    # Default filename if no path is provided
    if not excel_output_path:
        excel_output_path = './vehicles.xlsx'

    # Check if the file already exists
    file_exists = os.path.exists(excel_output_path)

    if file_exists:
        wb = load_workbook(excel_output_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        headers = ['rnr'] + keys
        ws.append(headers)

    # Insert data
    for vehicle in response_data['vehicle_data']:
        row = [vehicle.get('rnr', '')]
        for key in keys:
            row.append(vehicle.get(key, ''))
        ws.append(row)

        # Highlight rows based on `hu` values
        if colored and ws.max_row > 1:
            hu_date = vehicle.get('hu')
            if hu_date:
                hu_date = datetime.strptime(hu_date, '%Y-%m-%d')
                now = datetime.now()
                diff_months = (now.year - hu_date.year) * 12 + now.month - hu_date.month
                if diff_months <= 3:
                    fill = PatternFill(start_color="007500", end_color="007500", fill_type="solid")
                elif diff_months <= 12:
                    fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                else:
                    fill = PatternFill(start_color="b30000", end_color="b30000", fill_type="solid")

                for cell in ws[ws.max_row]:
                    cell.fill = fill

    # Save the workbook
    wb.save(excel_output_path)
    logging.info(f'File {excel_output_path} has been successfully created or updated.')


# Argument parser for command-line arguments
parser = argparse.ArgumentParser(description='Script to upload a CSV file and process the API response.')
parser.add_argument('-p', '--csv-path', type=str, required=True, help='Path to the CSV file')
parser.add_argument('-c', '--colored', action='store_true', default=True, help='Highlight rows based on conditions')
parser.add_argument('-o', '--excel-output-path', type=str, help='Optional path and filename for the Excel file')
parser.add_argument('-v', '--verbose', action='store_true', help='Enable verbose output')
args = parser.parse_args()

# Upload CSV file and process the response
response_data = upload_csv(args.csv_path, args.verbose)
process_and_save_to_excel(response_data, args.keys, args.colored, args.excel_output_path)

# python .\main.py -p C:/Users/nikolais/Documents/python-task/vehicles.csv -k rnr labelIds hu -c -v
# python .\main.py -p C:\Users\nicolai\PycharmProjects\python-task\vehicles.csv -k rnr labelIds hu -c -v